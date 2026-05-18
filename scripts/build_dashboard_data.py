from __future__ import annotations

import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
CURRENT_SALES_PATH = BASE_DIR / "data" / "current" / "Sales_vs_Stock_R5_Latest.xlsx"
SALES_DIR = BASE_DIR / "SALES"
REGION_SUMMARY_PATH = BASE_DIR / "ERA_ANALYTICS_Region5_Summary.xlsx"
OUTPUT_PATH = BASE_DIR / "outputs" / "dashboard_data_v2.json"
EAR_PNL_OUTPUT_PATH = BASE_DIR / "outputs" / "dashboard_data_ear_pnl.json"
IBOX_OUTPUT_PATH = BASE_DIR / "outputs" / "dashboard_data_ibox.json"
SAMSUNG_OUTPUT_PATH = BASE_DIR / "outputs" / "dashboard_data_samsung.json"

VALID_CHANNELS = {"ERAFONE", "ERA & MORE"}

# Toko yang sudah tutup — tidak akan pernah ditampilkan di dashboard
CLOSED_STORES: dict[str, set[str]] = {
    "IBOX": {"X108", "X056", "X019", "X013"},  # tutup per Mei 2026
    "SAMSUNG": set(),
    # EAR: toko yang tidak ada di LIST STORE TOKO 2026.xlsx (119 toko aktif)
    "EAR": {
        "E083", "E086", "E089", "E095", "E108", "E131", "E165", "E193",
        "E206", "E245", "E251", "E259", "E299", "E336", "E347", "E417",
        "E418", "E419", "E585", "E648", "E664", "E683", "E695", "E697",
        "E699", "E705", "E706", "E714", "E715", "E727", "E730", "E731",
        "E761", "E768", "E897", "E901", "E986", "E993", "F035", "F077",
        "F080", "F134", "M015", "M032", "M132", "M136", "M193", "M220",
    },
}

# Per-brand configuration
# summary_offset: column shift vs EAR base for all financial cols from col84 onwards
# analysis_remark_col: openpyxl col containing "This store is PROFIT/LOSS..." text
# analysis_total_col: None = sum cols 5-8; int = single "Count of Store" column
BRAND_CONFIG: dict[str, dict] = {
    "EAR": {
        "label": "Erafone (EAR)",
        "subfolder": "ERAFONE",
        "valid_categories": {"ERAFONE", "ERAFONE AND MORE"},
        "category_label": {"ERAFONE": "Erafone", "ERAFONE AND MORE": "ERA & More"},
        "summary_offset": 0,
        "analysis_remark_col": 3,
        "analysis_total_col": None,
    },
    "IBOX": {
        "label": "iBox (DCM)",
        "subfolder": "IBOX",
        "valid_categories": {"APP", "AAR", "APR"},
        "category_label": {"APP": "APP", "AAR": "AAR", "APR": "APR"},
        "summary_offset": 2,
        "analysis_remark_col": 2,
        "analysis_total_col": 19,
    },
    "SAMSUNG": {
        "label": "Samsung (NASA)",
        "subfolder": "SAMSUNG",
        "valid_categories": {"SES", "SEP", "SPS"},
        "category_label": {"SES": "SES", "SEP": "SEP", "SPS": "SPS"},
        "summary_offset": 0,
        "analysis_remark_col": 2,
        "analysis_total_col": 19,
    },
}

MONTH_ALIAS: dict[str, int] = {
    "JANUARY": 1, "FEBRUARY": 2, "MARCH": 3, "APRIL": 4,
    "MAY": 5, "JUNE": 6, "JULY": 7, "AUGUST": 8,
    "SEPTEMBER": 9, "OCTOBER": 10, "NOVEMBER": 11, "DECEMBER": 12,
    "JANUARI": 1, "FEBRUARI": 2, "MARET": 3,
    "MEI": 5, "JUNI": 6, "JULI": 7, "AGUSTUS": 8,
    "OKTOBER": 10, "DESEMBER": 12,
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4,
    "JUN": 6, "JUL": 7, "AUG": 8,
    "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}

MONTH_ABBR: dict[int, str] = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

MONTH_NAME_ID: dict[int, str] = {
    1: "Januari", 2: "Februari", 3: "Maret", 4: "April",
    5: "Mei", 6: "Juni", 7: "Juli", 8: "Agustus",
    9: "September", 10: "Oktober", 11: "November", 12: "Desember",
}

_PNL_FOLDER_RE = re.compile(r"^PNL EAR YTD (.+?)\s+(\d{4})$", re.IGNORECASE)


def resolve_pnl() -> tuple[Path, int, int]:
    """Return (pnl_folder, month_num, year) for the latest P&L period found."""
    candidates: list[tuple[int, Path, int, int]] = []
    for folder in BASE_DIR.iterdir():
        if not folder.is_dir():
            continue
        m = _PNL_FOLDER_RE.match(folder.name)
        if not m:
            continue
        month_str = m.group(1).strip().upper()
        year = int(m.group(2))
        month_num = MONTH_ALIAS.get(month_str)
        if not month_num:
            continue
        # Accept folder if it has at least one brand's xlsx (subfolder or root)
        has_data = (
            any((folder / cfg["subfolder"]).glob("*.xlsx") for cfg in BRAND_CONFIG.values())
            or any(
                f for f in folder.glob("*.xlsx")
                if "EAR" in f.name.upper()
                and "NASA" not in f.name.upper()
                and "DCM" not in f.name.upper()
            )
        )
        if not has_data:
            continue
        candidates.append((year * 12 + month_num, folder, month_num, year))

    if not candidates:
        raise FileNotFoundError(
            "Tidak ada folder P&L ditemukan. "
            "Buat folder dengan nama 'PNL EAR YTD <BULAN> <TAHUN>' di direktori proyek."
        )

    candidates.sort(key=lambda x: x[0], reverse=True)
    _, pnl_folder, month_num, year = candidates[0]
    return pnl_folder, month_num, year


def find_brand_file(pnl_folder: Path, brand_key: str) -> Path | None:
    """Find xlsx for a brand inside pnl_folder.

    New layout: pnl_folder/SUBFOLDER/*.xlsx
    Old layout: pnl_folder/*.xlsx filtered by filename keywords.
    """
    cfg = BRAND_CONFIG[brand_key]
    subfolder = pnl_folder / cfg["subfolder"]
    if subfolder.is_dir():
        files = list(subfolder.glob("*.xlsx"))
        return files[0] if files else None
    # Old/flat layout — filter by brand keyword
    keyword_map = {"EAR": ("EAR",), "IBOX": ("DCM",), "SAMSUNG": ("NASA",)}
    keywords = keyword_map.get(brand_key, ())
    files = [f for f in pnl_folder.glob("*.xlsx") if any(k in f.name.upper() for k in keywords)]
    return files[0] if files else None


def _pnl_sheets(sheet_names: list[str], month_num: int) -> dict[str, str]:
    """Detect old (Jan 2026) vs new (Feb 2026+) sheet layout."""
    if "R5" in sheet_names:
        return {
            "summary": "R5",
            "detail": "Report YTD",
            "analysis": "Analysis of Store EAR",
            "format": "old",
        }
    abbr = MONTH_ABBR[month_num]
    detail = f"YTD {abbr}"
    if detail not in sheet_names:
        ytd = [s for s in sheet_names if s.upper().startswith("YTD")]
        detail = ytd[0] if ytd else detail
    return {
        "summary": "Summary",
        "detail": detail,
        "analysis": "Analysis",
        "format": "new",
    }


def _latest_sales_in_folder() -> Path | None:
    if not SALES_DIR.exists():
        return None
    files = sorted(SALES_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    return files[0] if files else None


def resolve_sales_path() -> Path:
    if CURRENT_SALES_PATH.exists():
        return CURRENT_SALES_PATH
    latest = _latest_sales_in_folder()
    if latest:
        return latest
    raise FileNotFoundError("Tidak ada file Sales ditemukan. Letakkan file di folder SALES/ atau data/current/")


def compact_idr(value: float) -> str:
    abs_value = abs(value)
    if abs_value >= 1_000_000_000_000:
        return f"Rp{value / 1_000_000_000_000:.2f} T"
    if abs_value >= 1_000_000_000:
        return f"Rp{value / 1_000_000_000:.1f} M"
    if abs_value >= 1_000_000:
        return f"Rp{value / 1_000_000:.1f} jt"
    return f"Rp{value:,.0f}".replace(",", ".")


def month_total_columns(ws):
    cols = []
    for col in range(9, ws.max_column + 1):
        month_value = ws.cell(2, col).value
        sub_value = ws.cell(3, col).value
        if isinstance(month_value, datetime) and sub_value is None:
            cols.append((col, month_value))
    return cols


def build_sales_data():
    wb = load_workbook(resolve_sales_path(), read_only=True, data_only=True)
    ws = wb["BY STORE"]
    total_cols = month_total_columns(ws)
    latest_month = max((month for _, month in total_cols if month.year == 2026), default=None)
    latest_month_num = latest_month.month if latest_month else 12

    totals_by_year = defaultdict(float)
    stores = []

    for row in ws.iter_rows(min_row=4, values_only=True):
        channel = row[3]
        if channel not in VALID_CHANNELS:
            continue

        monthly = {}
        for col, month_dt in total_cols:
            value = row[col - 1] or 0
            key = month_dt.strftime("%Y-%m")
            monthly[key] = float(value)
            totals_by_year[month_dt.year] += float(value)

        yearly = {
            "y2022": sum(monthly.get(f"2022-{m:02d}", 0) for m in range(1, 13)),
            "y2023": sum(monthly.get(f"2023-{m:02d}", 0) for m in range(1, 13)),
            "y2024": sum(monthly.get(f"2024-{m:02d}", 0) for m in range(1, 13)),
            "y2025": sum(monthly.get(f"2025-{m:02d}", 0) for m in range(1, 13)),
            "y2026": sum(monthly.get(f"2026-{m:02d}", 0) for m in range(1, latest_month_num + 1)),
        }

        q125 = sum(monthly.get(f"2025-{m:02d}", 0) for m in range(1, latest_month_num + 1))
        q126 = sum(monthly.get(f"2026-{m:02d}", 0) for m in range(1, latest_month_num + 1))
        latest_period_sales = monthly.get(latest_month.strftime("%Y-%m"), 0) if latest_month else 0
        bep = float(row[1] or 0)
        yoy = ((q126 / q125) - 1) if q125 else 0
        bep_ach = (latest_period_sales / bep) if bep else 0
        status = row[0] or "Unknown"
        tsh = row[4] or "Unknown"

        if status != "Aktif":
            cluster = "Recovery"
        elif bep_ach >= 1 and yoy >= 0.08:
            cluster = "Growth"
        elif bep_ach >= 1:
            cluster = "Stable"
        elif bep_ach >= 0.7:
            cluster = "Recovery"
        else:
            cluster = "Critical"

        health = max(
            30,
            min(
                92,
                round(
                    42
                    + (min(bep_ach, 1.4) * 18)
                    + (max(min(yoy, 0.45), -0.3) * 22)
                    + (5 if status == "Aktif" else -10)
                ),
            ),
        )

        stores.append(
            {
                "code": row[6],
                "name": row[7],
                "channel": "Erafone" if channel == "ERAFONE" else "ERA & More",
                "tsh": tsh,
                "status": status,
                "cluster": cluster,
                "health": health,
                "sales": yearly,
                "q125": q125,
                "q126": q126,
                "yoy": yoy * 100,
                "mar26": latest_period_sales,
                "latestPeriodSales": latest_period_sales,
                "latestPeriodLabel": latest_month.strftime("%b %Y") if latest_month else "Latest",
                "bep": bep,
                "bep_ach": bep_ach,
                "annualTotal": yearly["y2022"] + yearly["y2023"] + yearly["y2024"] + yearly["y2025"] + yearly["y2026"],
                "bepGap": latest_period_sales - bep if bep else 0,
            }
        )

    return {
        "stores": stores,
        "yearTotals": {str(year): value for year, value in sorted(totals_by_year.items())},
        "latestMonthLabel": latest_month.strftime("%b %Y") if latest_month else "Latest",
        "latestMonthNumber": latest_month_num,
    }


def parse_r5_summary(wb, sheets: dict, month_num: int, year: int, cfg: dict | None = None) -> dict:
    cfg = cfg or BRAND_CONFIG["EAR"]
    ws = wb[sheets["summary"]]
    is_old = sheets["format"] == "old"
    offset = cfg["summary_offset"]

    # Old format (Jan 2026) EAR-only column positions; new format uses base + offset
    if is_old:
        ho_finance_col, net_final_col = 116, 118
        erafone_row, eam_row = 11, 21
        opex_col, oi_col = 84, 86
        net_before_ho, ho_cost, net_after_ho = 97, 104, 106
        finance_cost, net_after_finance = 110, 112
    else:
        opex_col = 84 + offset
        oi_col = 86 + offset
        net_before_ho = 97 + offset
        ho_cost = 104 + offset
        net_after_ho = 106 + offset
        finance_cost = 110 + offset
        net_after_finance = 112 + offset
        ho_finance_col = 122 + offset
        net_final_col = 124 + offset
        erafone_row, eam_row = 10, 20

    grand = {
        "netSales": ws.cell(5, 11).value or 0,
        "grossProfit": ws.cell(5, 15).value or 0,
        "gpPct": ws.cell(5, 16).value or 0,
        "totalOpex": ws.cell(5, opex_col).value or 0,
        "operatingIncome": ws.cell(5, oi_col).value or 0,
        "netBeforeHo": ws.cell(5, net_before_ho).value or 0,
        "hoCost": ws.cell(5, ho_cost).value or 0,
        "netAfterHo": ws.cell(5, net_after_ho).value or 0,
        "financeCost": ws.cell(5, finance_cost).value or 0,
        "netAfterFinance": ws.cell(5, net_after_finance).value or 0,
        "hoFinanceCost": ws.cell(5, ho_finance_col).value or 0,
        "netFinal": ws.cell(5, net_final_col).value or 0,
    }

    cost_drivers = [
        ("Selling salary", ws.cell(5, 21).value or 0),
        ("ROU depreciation", ws.cell(5, 19).value or 0),
        ("Credit card", ws.cell(5, 23).value or 0),
        ("Fixed asset dep.", ws.cell(5, 43).value or 0),
    ]

    return {
        "periodLabel": f"Profitability Check YTD {MONTH_NAME_ID[month_num]} {year}",
        "grand": grand,
        "erafoneNetFinal": ws.cell(erafone_row, net_final_col).value or 0,
        "eraMoreNetFinal": ws.cell(eam_row, net_final_col).value or 0,
        "costDrivers": cost_drivers,
    }


def parse_analysis_of_store_ear(wb, sheets: dict, cfg: dict | None = None) -> dict:
    ws = wb[sheets["analysis"]]

    if sheets["format"] == "old":
        # Old "Analysis of Store EAR": larger sheet, different column layout
        # Col 1=rank, col 2=remark, col 3=mall, col 4=street, col 5=total
        rows = []
        for row_idx in range(4, 140):
            remark = str(ws.cell(row_idx, 2).value or "").strip()
            if not remark or remark.upper() in {"REMARK", "GRAND TOTAL"}:
                continue
            mall = int(ws.cell(row_idx, 3).value or 0)
            street = int(ws.cell(row_idx, 4).value or 0)
            total = int(ws.cell(row_idx, 5).value or 0)
            rows.append({
                "rank": str(ws.cell(row_idx, 1).value or "").strip(),
                "remark": remark,
                "regionTotal": total,
                "mallTotal": mall,
                "streetTotal": street,
            })
            if not rows[-1]["rank"] and not remark:
                rows.pop()
        # Find grand total row
        grand_total = sum(r["regionTotal"] for r in rows)
        grand = {
            "regionTotal": grand_total, "mallTotal": 0, "streetTotal": 0,
            "erafoneTotal": grand_total, "eraMoreTotal": 0,
        }
    else:
        # New format: remark col and total col differ per brand
        cfg = cfg or BRAND_CONFIG["EAR"]
        remark_col = cfg["analysis_remark_col"]
        total_col = cfg["analysis_total_col"]  # None = sum cols 5-8; int = single col

        def _row_total(row_idx: int) -> int:
            if total_col is not None:
                return int(ws.cell(row_idx, total_col).value or 0)
            return sum(int(ws.cell(row_idx, c).value or 0) for c in range(5, 9))

        rows = []
        for row_idx in range(4, ws.max_row):
            remark = str(ws.cell(row_idx, remark_col).value or "").strip()
            if not remark or remark == "Grand Total":
                continue
            rows.append({
                "rank": str(ws.cell(row_idx, 1).value or "").strip(),
                "remark": remark,
                "regionTotal": _row_total(row_idx),
                "mallTotal": 0,
                "streetTotal": 0,
            })
        grand_row = ws.max_row
        grand_total = _row_total(grand_row)
        grand = {
            "regionTotal": grand_total, "mallTotal": 0, "streetTotal": 0,
            "erafoneTotal": grand_total, "eraMoreTotal": 0,
        }

    top_profit = max((r for r in rows if "PROFIT" in r["remark"]), key=lambda x: x["regionTotal"], default=None)
    top_loss = max((r for r in rows if "LOSS" in r["remark"]), key=lambda x: x["regionTotal"], default=None)

    return {
        "regionLabel": "REGION 5",
        "grand": grand,
        "topProfit": top_profit,
        "topLoss": top_loss,
        "rows": rows,
    }


def parse_pnl_store_details(wb, sheets: dict, cfg: dict | None = None) -> dict:
    cfg = cfg or BRAND_CONFIG["EAR"]
    ws = wb[sheets["detail"]]
    is_old = sheets["format"] == "old"

    min_row = 6 if is_old else 5
    opex_col = 84 if is_old else 90
    oi_col = 86 if is_old else 92
    ho_fin_col = 119 if is_old else 125
    net_final_col = 121 if is_old else 127

    valid_categories = cfg["valid_categories"]
    category_label = cfg["category_label"]

    mapping = {}
    for row in ws.iter_rows(min_row=min_row, values_only=True):
        region = row[6]
        category = row[4]
        if region != "REGION 5" or category not in valid_categories:
            continue
        code = row[0]
        if not code:
            continue
        mapping[code] = {
            "name": str(row[1] or "").strip(),
            "pnlCategory": category_label.get(category, category),
            "netSalesPnl": row[17] or 0,
            "grossProfitPnl": row[21] or 0,
            "totalOpexPnl": row[opex_col] or 0,
            "operatingIncomePnl": row[oi_col] or 0,
            "netBeforeHo": row[103] or 0,
            "hoCost": row[107] or 0,
            "netAfterHo": row[109] or 0,
            "financeCost": row[113] or 0,
            "netAfterFinance": row[115] or 0,
            "hoFinanceCost": row[ho_fin_col] or 0,
            "netFinal": row[net_final_col] or 0,
        }

    return mapping


def parse_region_summary():
    wb = load_workbook(REGION_SUMMARY_PATH, read_only=True, data_only=True)

    ws_tsh = wb["1. TSH Summary"]
    tsh_summary = {}
    meta = {"grandTotalStores": 0, "annualBep": 0, "annualRugi": 0, "grand2025": 0, "active2025": 0}
    excluded_rows = {"Close", "GRAND TOTAL", "YoY Growth"}

    for row in ws_tsh.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        name = str(row[0]).strip()
        if name == "GRAND TOTAL":
            meta["grand2025"] = float(row[4] or 0)
            continue
        if name in excluded_rows:
            continue
        tsh_summary[name] = {
            "tsh": name,
            "y2022": float(row[1] or 0),
            "y2023": float(row[2] or 0),
            "y2024": float(row[3] or 0),
            "y2025": float(row[4] or 0),
            "y2026": float(row[5] or 0),
            "annualTotal": float(row[6] or 0),
        }
        meta["active2025"] += float(row[4] or 0)

    ws_bep = wb["2. BEP Status"]
    bep_status_map = {}
    for row in ws_bep.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        code = str(row[1]).strip()
        status = str(row[7] or "").strip()
        bep_status_map[code] = {
            "bepTarget": float(row[5] or 0),
            "bestMonthSales": float(row[6] or 0),
            "annualBepStatus": status,
        }
        if status == "BEP REACHED":
            meta["annualBep"] += 1
        elif status == "BELOW BEP":
            meta["annualRugi"] += 1

    ws_store = wb["3. Store Annual"]
    store_summary_map = {}
    for row in ws_store.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        code = str(row[1]).strip()
        store_summary_map[code] = {
            "code": code,
            "name": row[2],
            "channel": "Erafone" if row[3] == "ERAFONE" else "ERA & More",
            "tsh": row[4] or "Unknown",
            "sales": {
                "y2022": float(row[5] or 0),
                "y2023": float(row[6] or 0),
                "y2024": float(row[7] or 0),
                "y2025": float(row[8] or 0),
                "y2026": float(row[9] or 0),
            },
            "annualTotal": float(row[10] or 0),
            "annualStatus": str(row[11] or "").strip(),
        }
        meta["grandTotalStores"] += 1

    return {
        "tshSummary": tsh_summary,
        "storeSummaryMap": store_summary_map,
        "bepStatusMap": bep_status_map,
        "meta": meta,
    }


def enrich_stores(sales_stores, pnl_map, region_summary):
    enriched = []
    store_summary_map = region_summary["storeSummaryMap"]
    bep_status_map = region_summary["bepStatusMap"]
    for store in sales_stores:
        summary_store = store_summary_map.get(store["code"], {})
        if summary_store:
            store["name"] = summary_store["name"] or store["name"]
            store["channel"] = summary_store["channel"] or store["channel"]
            store["tsh"] = summary_store["tsh"] or store["tsh"]
            store["sales"] = summary_store["sales"]
            store["annualTotal"] = summary_store["annualTotal"]
            store["annualStatus"] = summary_store["annualStatus"]
        else:
            store["annualStatus"] = "N/A"

        annual_bep = bep_status_map.get(store["code"], {})
        store["annualBepStatus"] = annual_bep.get("annualBepStatus", "N/A")
        store["bestMonthSales"] = float(annual_bep.get("bestMonthSales", 0))
        if annual_bep.get("bepTarget"):
            store["bep"] = float(annual_bep["bepTarget"])

        pnl = pnl_map.get(store["code"], {})
        net_final = float(pnl.get("netFinal", 0))
        store["pnl"] = "Profit" if net_final > 0 else "Loss" if net_final < 0 else "Flat"
        store["netFinal"] = net_final
        store["operatingIncomePnl"] = float(pnl.get("operatingIncomePnl", 0))

        if store["tsh"] == "VACANT":
            action = "Assign owner + review cost"
        elif store["cluster"] == "Critical" and store["bepGap"] < -1_000_000_000:
            action = "Traffic + cost review"
        elif store["cluster"] == "Critical":
            action = "Conversion + bundle"
        elif store["cluster"] == "Recovery":
            action = "Push recovery plan"
        elif store["cluster"] == "Growth":
            action = "Scale best practice"
        else:
            action = "Maintain performance"

        store["action"] = action
        enriched.append(store)
    return enriched


def build_tsh_stats(stores, tsh_summary):
    active_operational = defaultdict(lambda: {"stores": 0, "bepOperational": 0, "lossOperational": 0, "avgHealth": 0})
    for store in stores:
        if store["status"] != "Aktif":
            continue
        tsh = store["tsh"]
        active_operational[tsh]["stores"] += 1
        active_operational[tsh]["bepOperational"] += 1 if store["bepGap"] >= 0 else 0
        active_operational[tsh]["lossOperational"] += 1 if store["pnl"] == "Loss" else 0
        active_operational[tsh]["avgHealth"] += store["health"]

    result = []
    for tsh, summary in tsh_summary.items():
        active = active_operational.get(tsh, {})
        active_stores = active.get("stores", 0)
        annual_stores = sum(1 for s in stores if s["tsh"] == tsh)
        annual_bep = sum(1 for s in stores if s["tsh"] == tsh and s.get("annualStatus") == "BEP")
        annual_loss = sum(1 for s in stores if s["tsh"] == tsh and s.get("annualStatus") == "RUGI")
        result.append(
            {
                "tsh": tsh,
                "stores": annual_stores,
                "activeStores": active_stores,
                "bep": annual_bep,
                "loss": annual_loss,
                "bepOperational": active.get("bepOperational", 0),
                "lossOperational": active.get("lossOperational", 0),
                "total2025": summary["y2025"],
                "total2026": summary["y2026"],
                "avgHealth": round(active.get("avgHealth", 0) / active_stores, 1) if active_stores else 0,
            }
        )
    return sorted(result, key=lambda x: x["total2025"], reverse=True)


def build_actions(stores, tsh_stats, analysis_summary):
    critical = sorted(
        [s for s in stores if s["cluster"] == "Critical" and s["status"] == "Aktif"],
        key=lambda x: x["bepGap"],
    )[:25]
    biggest_gap = critical[0] if critical else None
    worst_tsh = sorted(tsh_stats, key=lambda x: x["loss"], reverse=True)[0] if tsh_stats else None
    top_loss = analysis_summary.get("topLoss")
    return [
        {
            "title": f"Review {len(critical)} toko critical",
            "note": "Prioritaskan store yang gap BEP paling besar dan P&L final masih negatif.",
            "tag": "Urgent",
        },
        {
            "title": f"{biggest_gap['name']} - gap {compact_idr(abs(biggest_gap['bepGap']))}" if biggest_gap else "Gap terbesar",
            "note": "Store dengan jarak paling jauh dari target BEP bulanannya.",
            "tag": "Kritis",
        },
        {
            "title": f"{top_loss['regionTotal']} toko masuk loss pattern utama" if top_loss else f"{sum(1 for s in stores if s['bepGap'] >= 0)} toko sudah capai BEP",
            "note": top_loss["remark"].replace("This store is ", "") if top_loss else "Gunakan toko sehat sebagai benchmark bundle, attach, dan cadangan best practice.",
            "tag": "Info",
        },
        {
            "title": f"{worst_tsh['tsh']} - {worst_tsh['loss']} toko loss" if worst_tsh else "Watchlist TSH",
            "note": "Owner dengan toko rugi terbanyak perlu weekly review khusus.",
            "tag": "Review",
        },
    ]


def build_executive(stores, sales_year_totals, tsh_stats, region_meta, analysis_summary, sales_meta):
    active = [s for s in stores if s["status"] == "Aktif"]
    growth = sum(1 for s in stores if s["cluster"] == "Growth" and s["status"] == "Aktif")
    stable = sum(1 for s in stores if s["cluster"] == "Stable" and s["status"] == "Aktif")
    recovery = sum(1 for s in stores if s["cluster"] == "Recovery" and s["status"] == "Aktif")
    critical = sum(1 for s in stores if s["cluster"] == "Critical" and s["status"] == "Aktif")
    avg_health = round(sum(s["health"] for s in active) / len(active), 1) if active else 0
    yoy = (
        (
            sum(s["q126"] for s in stores if s["status"] == "Aktif")
            / max(sum(s["q125"] for s in stores if s["status"] == "Aktif"), 1)
        )
        - 1
    ) * 100
    critical_tsh = sorted(tsh_stats, key=lambda x: x["loss"], reverse=True)[0]["tsh"] if tsh_stats else "-"
    top_profit = analysis_summary.get("topProfit")
    top_loss = analysis_summary.get("topLoss")

    summary_title = "Good - On Track Recovery" if critical < recovery + growth else "Mixed - Turnaround Needed"
    condition = f"{growth + stable} dari {len(active)} toko aktif masuk Growth/Stable. Fokus review ada pada {critical} toko critical dengan gap BEP terbesar."
    analysis = f"TSH {critical_tsh} memiliki konsentrasi toko loss tertinggi. Store format besar dan owner vacant menjadi sinyal utama root cause operasional."
    action = "Mulai dari cluster Critical, validasi cost structure, dorong traffic lokal, lalu replikasi best practice dari store Growth."
    if top_profit and top_loss:
        condition = (
            f"{analysis_summary['grand']['regionTotal']} toko Region 5 terbaca di Analysis EAR. "
            f"Loss pattern terbesar ada di {top_loss['regionTotal']} toko, sedangkan profit pattern terbesar ada di {top_profit['regionTotal']} toko."
        )
        analysis = (
            f"Pattern dominan profit: {top_profit['remark'].replace('This store is ', '')}. "
            f"Pattern dominan loss: {top_loss['remark'].replace('This store is ', '')}."
        )
        action = "Prioritaskan toko yang masuk loss pattern terbesar, terutama yang sales masih below dan OPEX masih above, lalu tindak lanjutkan per owner dan format toko."

    return {
        "totalStores": len(stores),
        "activeStores": len(active),
        "bepReached": growth + stable,
        "belowBep": recovery + critical,
        "annualBepReached": region_meta["annualBep"],
        "annualBelowBep": region_meta["annualRugi"],
        "closedStores": len(stores) - len(active),
        "revenue2025": compact_idr(region_meta["active2025"] or sales_year_totals.get("2025", 0)),
        "yoy": round(yoy, 1),
        "avgHealth": avg_health,
        "dataPeriod": sales_meta["latestMonthLabel"],
        "dataRange": "2022 - 2026",
        "totalStoresNote": f"{len(active)} aktif + {len(stores) - len(active)} close (summary 2022-2026)",
        "bepReachedNote": f"{growth + stable} Growth/Stable aktif · annual BEP {region_meta['annualBep']}",
        "belowBepNote": f"{critical + recovery} recovery/critical aktif · annual rugi {region_meta['annualRugi']}",
        "revenue2025Note": "Historical annual resmi dari workbook summary",
        "yoyNote": f"Jan-{sales_meta['latestMonthLabel'].split()[0]} 2026 vs Jan-{sales_meta['latestMonthLabel'].split()[0]} 2025",
        "summarySubtitle": "Hybrid view: REPORT YTD + Analysis of Store EAR REGION 5 + sales operational",
        "clusters": {"Growth": growth, "Stable": stable, "Recovery": recovery, "Critical": critical},
        "summaryTitle": summary_title,
        "condition": condition,
        "analysis": analysis,
        "action": action,
    }


def build_brand_pnl_only(brand_key: str, path: Path, month_num: int, year: int) -> dict:
    """Build a P&L-only dashboard payload for IBOX or SAMSUNG (no sales workbook needed)."""
    cfg = BRAND_CONFIG[brand_key]
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = _pnl_sheets(wb.sheetnames, month_num)

    pnl_summary = parse_r5_summary(wb, sheets, month_num, year, cfg)
    analysis_summary = parse_analysis_of_store_ear(wb, sheets, cfg)
    store_map = parse_pnl_store_details(wb, sheets, cfg)
    wb.close()

    closed = CLOSED_STORES.get(brand_key, set())
    stores = []
    for code, pnl in store_map.items():
        if code in closed:
            continue
        net_final = float(pnl["netFinal"])
        pnl_status = "Profit" if net_final > 0 else "Loss" if net_final < 0 else "Flat"
        stores.append({
            "code": code,
            "name": pnl.get("name", code),
            "pnlCategory": pnl["pnlCategory"],
            "pnl": pnl_status,
            "netFinal": net_final,
            "netSalesPnl": float(pnl["netSalesPnl"]),
            "grossProfitPnl": float(pnl["grossProfitPnl"]),
            "totalOpexPnl": float(pnl["totalOpexPnl"]),
            "operatingIncomePnl": float(pnl["operatingIncomePnl"]),
            "hoFinanceCost": float(pnl["hoFinanceCost"]),
        })

    profit_stores = [s for s in stores if s["pnl"] == "Profit"]
    loss_stores = [s for s in stores if s["pnl"] == "Loss"]
    grand = pnl_summary["grand"]

    executive = {
        "brand": cfg["label"],
        "totalStores": len(stores),
        "profitStores": len(profit_stores),
        "lossStores": len(loss_stores),
        "totalNetSales": grand["netSales"],
        "totalGrossProfit": grand["grossProfit"],
        "gpPct": grand["gpPct"],
        "totalOpex": grand["totalOpex"],
        "operatingIncome": grand["operatingIncome"],
        "netFinal": grand["netFinal"],
        "periodLabel": pnl_summary["periodLabel"],
        "summaryNetSales": compact_idr(grand["netSales"]),
        "summaryNetFinal": compact_idr(grand["netFinal"]),
    }

    pnl_summary["analysis"] = analysis_summary

    return {
        "brand": brand_key,
        "executive": executive,
        "stores": sorted(stores, key=lambda s: s["netFinal"], reverse=True),
        "pnl": pnl_summary,
        "source": {
            "pnlWorkbook": str(path.relative_to(BASE_DIR)),
            "pnlPeriod": f"{MONTH_NAME_ID[month_num]} {year}",
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
        },
    }


def build_dashboard_data():
    pnl_folder, month_num, year = resolve_pnl()
    period_label = f"{MONTH_NAME_ID[month_num]} {year}"
    print(f"P&L folder: {pnl_folder.name} | period: {period_label}")

    # ── EAR (full pipeline with sales) ──────────────────────────────────────
    ear_path = find_brand_file(pnl_folder, "EAR")
    if not ear_path:
        raise FileNotFoundError("File P&L EAR tidak ditemukan di subfolder ERAFONE/")

    wb_pnl = load_workbook(ear_path, read_only=True, data_only=True)
    sheets = _pnl_sheets(wb_pnl.sheetnames, month_num)
    print(f"  EAR: {ear_path.name} | format={sheets['format']} | sheets: {sheets['summary']}, {sheets['detail']}, {sheets['analysis']}")

    sales = build_sales_data()
    ear_cfg = BRAND_CONFIG["EAR"]
    pnl_summary = parse_r5_summary(wb_pnl, sheets, month_num, year, ear_cfg)
    analysis_summary = parse_analysis_of_store_ear(wb_pnl, sheets, ear_cfg)
    pnl_store_map = parse_pnl_store_details(wb_pnl, sheets, ear_cfg)
    wb_pnl.close()

    region_summary = parse_region_summary()
    stores = enrich_stores(sales["stores"], pnl_store_map, region_summary)
    tsh_stats = build_tsh_stats(stores, region_summary["tshSummary"])
    actions = build_actions(stores, tsh_stats, analysis_summary)
    executive = build_executive(stores, sales["yearTotals"], tsh_stats, region_summary["meta"], analysis_summary, sales)
    pnl_summary["analysis"] = analysis_summary
    top_critical = sorted([s for s in stores if s["cluster"] == "Critical"], key=lambda x: x["bepGap"])[:6]
    sales_path = resolve_sales_path()

    ear_data = {
        "executive": executive,
        "actions": actions,
        "stores": stores,
        "tshStats": tsh_stats,
        "topCritical": top_critical,
        "pnl": pnl_summary,
        "source": {
            "salesWorkbook": str(sales_path.relative_to(BASE_DIR)),
            "salesWorkbookActive": str(sales_path.relative_to(BASE_DIR)),
            "pnlWorkbook": str(ear_path.relative_to(BASE_DIR)),
            "pnlPeriod": period_label,
            "summaryWorkbook": str(REGION_SUMMARY_PATH.relative_to(BASE_DIR)),
            "generatedAt": datetime.now().isoformat(timespec="seconds"),
        },
    }

    # ── EAR P&L-only (sama struktur dengan iBox/Samsung, tanpa data sales) ──
    print(f"  EAR P&L-only: {ear_path.name}")
    ear_pnl_data = build_brand_pnl_only("EAR", ear_path, month_num, year)

    # ── IBOX ────────────────────────────────────────────────────────────────
    ibox_path = find_brand_file(pnl_folder, "IBOX")
    ibox_data = None
    if ibox_path:
        print(f"  IBOX: {ibox_path.name}")
        ibox_data = build_brand_pnl_only("IBOX", ibox_path, month_num, year)

    # ── SAMSUNG ─────────────────────────────────────────────────────────────
    samsung_path = find_brand_file(pnl_folder, "SAMSUNG")
    samsung_data = None
    if samsung_path:
        print(f"  SAMSUNG: {samsung_path.name}")
        samsung_data = build_brand_pnl_only("SAMSUNG", samsung_path, month_num, year)

    return ear_data, ear_pnl_data, ibox_data, samsung_data


def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    ear_data, ear_pnl_data, ibox_data, samsung_data = build_dashboard_data()

    OUTPUT_PATH.write_text(json.dumps(ear_data, ensure_ascii=False, indent=2))
    print(f"  → {OUTPUT_PATH.name} ({len(ear_data['stores'])} stores EAR full)")

    EAR_PNL_OUTPUT_PATH.write_text(json.dumps(ear_pnl_data, ensure_ascii=False, indent=2))
    ex = ear_pnl_data["executive"]
    print(f"  → {EAR_PNL_OUTPUT_PATH.name} ({ex['totalStores']} stores EAR | {ex['profitStores']} Profit / {ex['lossStores']} Loss)")

    if ibox_data:
        IBOX_OUTPUT_PATH.write_text(json.dumps(ibox_data, ensure_ascii=False, indent=2))
        print(f"  → {IBOX_OUTPUT_PATH.name} ({len(ibox_data['stores'])} stores iBox)")

    if samsung_data:
        SAMSUNG_OUTPUT_PATH.write_text(json.dumps(samsung_data, ensure_ascii=False, indent=2))
        print(f"  → {SAMSUNG_OUTPUT_PATH.name} ({len(samsung_data['stores'])} stores Samsung)")


if __name__ == "__main__":
    main()
