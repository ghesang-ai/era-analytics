from __future__ import annotations

import argparse
import shutil
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
CURRENT_DIR = BASE_DIR / "data" / "current"
ARCHIVE_DIR = BASE_DIR / "data" / "archive"
TARGET_PATH = CURRENT_DIR / "Sales_vs_Stock_R5_Latest.xlsx"
DEFAULT_DOWNLOADS = Path.home() / "Downloads"


def find_latest_download() -> Path | None:
    matches = sorted(
        DEFAULT_DOWNLOADS.glob("Sales vs Stock * R5_Master.xlsx"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return matches[0] if matches else None


def validate_workbook(path: Path) -> None:
    wb = load_workbook(path, read_only=True, data_only=True)
    if "BY STORE" not in wb.sheetnames:
        raise ValueError("Sheet 'BY STORE' tidak ditemukan.")
    ws = wb["BY STORE"]
    headers = [ws.cell(2, col).value for col in range(1, 9)]
    expected = ["Status", "BEP", "KET SSSG", "Channel", "TSH", "TSH YTD", "Code", "Store Name"]
    if headers != expected:
        raise ValueError(f"Header BY STORE berubah. Ditemukan: {headers}")


def archive_previous() -> None:
    if not TARGET_PATH.exists():
        return
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archived = ARCHIVE_DIR / f"Sales_vs_Stock_R5_{timestamp}.xlsx"
    shutil.copy2(TARGET_PATH, archived)


def main() -> None:
    parser = argparse.ArgumentParser(description="Import latest daily Sales vs Stock workbook into ERA-ANALYTICS V1.")
    parser.add_argument("source", nargs="?", help="Path file Sales vs Stock harian.")
    args = parser.parse_args()

    source = Path(args.source).expanduser() if args.source else find_latest_download()
    if not source or not source.exists():
        raise SystemExit("File harian tidak ditemukan. Berikan path file atau letakkan file di Downloads.")

    validate_workbook(source)
    CURRENT_DIR.mkdir(parents=True, exist_ok=True)
    archive_previous()
    shutil.copy2(source, TARGET_PATH)

    subprocess.run(
        ["python3", str(BASE_DIR / "scripts" / "build_dashboard_data.py")],
        cwd=BASE_DIR,
        check=True,
    )

    print(f"Imported: {source}")
    print(f"Active source: {TARGET_PATH}")


if __name__ == "__main__":
    main()
